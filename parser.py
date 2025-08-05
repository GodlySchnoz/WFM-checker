import re
import requests
import openpyxl
import csv
import argparse
import sys
from datetime import datetime, timezone

def parse_input(file_path):
    """Main input parser that routes to format-specific parsers"""
    if file_path.endswith('.txt'):
        return parse_txt(file_path)
    elif file_path.endswith('.csv'):
        return parse_csv(file_path)
    elif file_path.endswith('.xlsx'):
        return parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path}")

def parse_txt(file_path):
    """Text file parser with support for optional rank parameter"""
    items = []
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    for line in lines:
        line = line.strip()
        if not line or line.endswith(':'):
            continue
        # Remove category labels
        line = re.sub(r'^[\w\s]+:\s*', '', line)

        # Split into individual entries
        entries = re.split(r',\s*', line)

        for entry in entries:
            entry = entry.strip()
            if not entry:
                continue

            # Match patterns with quantity and optional rank
            match = re.match(r'^(\d+)\s+(?:copies of|copy of|of)?\s*(.+?)(?:\s+(\d+))?$', entry, re.IGNORECASE)
            if match:
                quantity = int(match.group(1))
                item_name = match.group(2).strip().lower()
                rank = int(match.group(3) or 0)  # Default to 0 if not provided
                items.append((quantity, item_name, rank))
            else:
                # Handle entries without explicit quantity
                tokens = entry.split()
                if tokens and tokens[-1].isdigit():
                    rank = int(tokens[-1])
                    item_name = " ".join(tokens[:-1]).lower()
                    items.append((1, item_name, rank))
                else:
                    items.append((1, entry.lower(), 0))
    return items

def parse_csv(file_path):
    """CSV parser with support for optional rank column"""
    items = []
    with open(file_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        headers = next(reader, None)
        
        # Detect column positions
        quantity_col, item_col, rank_col = detect_columns(headers)
        
        for row in reader:
            if len(row) <= max(quantity_col, item_col):
                continue
                
            quantity = parse_quantity(row[quantity_col])
            item_name = row[item_col].strip().lower()
            
            # Get rank if available, default to 0
            rank = 0
            if rank_col is not None and len(row) > rank_col:
                try:
                    rank = int(row[rank_col])
                except (ValueError, TypeError):
                    rank = 0
            
            if item_name:
                items.append((quantity, item_name, rank))
    return items

def parse_xlsx(file_path):
    """Excel parser with support for optional rank column"""
    items = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Detect column positions
    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    quantity_col, item_col, rank_col = detect_columns(headers)
    
    start_row = 2 if headers else 1
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= max(quantity_col, item_col):
            continue
        if row[item_col] is None:
            continue
        quantity = parse_quantity(row[quantity_col])
        item_name = str(row[item_col]).strip().lower()
        
        # Get rank if available, default to 0
        rank = 0
        if rank_col is not None and len(row) > rank_col and row[rank_col] is not None:
            try:
                rank = int(row[rank_col])
            except (ValueError, TypeError):
                rank = 0
            
        if item_name:
            items.append((quantity, item_name, rank))
    return items

def detect_columns(headers):
    """Detect quantity, item, and rank columns"""
    quantity_col = 0
    item_col = 1
    rank_col = None
    
    if headers:
        header_lower = [str(h).lower().strip() for h in headers]
        try:
            quantity_col = header_lower.index('quantity')
        except ValueError:
            pass
        try:
            item_col = header_lower.index('item')
        except ValueError:
            pass
        try:
            rank_col = header_lower.index('rank')
        except ValueError:
            pass
    return quantity_col, item_col, rank_col

def parse_quantity(value):
    """Parse quantity from various input types"""
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 1

def normalize_item_name(item_name):
    """
    Normalizes the item name to match the Warframe Market API's expected format.
    """

    item_name = item_name.lower().strip()

    if item_name in special_cases:
        return special_cases[item_name]

    # Normalizer for most items
    item_name = (
        item_name
        .replace("&", "and")
        .replace(".", "")
        .replace("-", "_")
        .replace(" ", "_")
        .replace("'", "")
        .replace("orokin", "corrupted")
    )

    # Check item suffixes and add blueprint if necessary
    if item_name.endswith(('_systems', '_chassis', "_harness", "_wings")) and item_name not in exceptions:
        item_name += '_blueprint'

    return item_name

def get_ayatan_star_data(item_name):
    """
    Returns the star counts for Ayatan sculptures.
    Returns (amber_stars, cyan_stars) tuple or (None, None) if not an Ayatan sculpture.
    """
    normalized_name = normalize_item_name(item_name)
    
    if normalized_name in ayatan_sculptures:
        return ayatan_sculptures[normalized_name]
    
    return None, None  # Not an Ayatan sculpture


def get_item_price_stat(item_name, price_method, rank):
    """
    Fetches the most recent closed price from the statistics API.
    """
    url_name = normalize_item_name(item_name)
    api_url = f"https://api.warframe.market/v1/items/{url_name}/statistics"

    headers = {
        'accept': 'application/json',
        'platform': 'pc',
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()
        data = response.json()

        # Get the statistics_closed data
        if 'payload' not in data or 'statistics_closed' not in data['payload']:
            print(f"No statistics data available for '{item_name}'")
            return None
            
        stats_closed = data['payload']['statistics_closed']
        
        # Get the 48hours data (most recent timeframe)
        if '48hours' not in stats_closed or not stats_closed['48hours']: #TODO add handling for missing 48hours data
            print(f"No 48-hour statistics available for '{item_name}'")
            return None
            
        closed_data = stats_closed['48hours']
        
        # Filter by rank/stars if specified
        has_mod_rank_data = any('mod_rank' in entry for entry in closed_data)
        has_cyan_stars_data = any('cyan_stars' in entry for entry in closed_data)
        has_amber_stars_data = any('amber_stars' in entry for entry in closed_data)
        
        if has_cyan_stars_data or has_amber_stars_data:
            # Check if this is an Ayatan sculpture with specific star requirements
            target_amber, target_cyan = get_ayatan_star_data(item_name)
            
            if target_amber is not None and target_cyan is not None:
                # Use specific Ayatan sculpture star data
                star_filtered_data = []
                for entry in closed_data:
                    cyan_stars = entry.get('cyan_stars', 0)
                    amber_stars = entry.get('amber_stars', 0)
                    if cyan_stars == target_cyan and amber_stars == target_amber:
                        star_filtered_data.append(entry)

                if star_filtered_data:
                    closed_data = star_filtered_data
                else:
                    # If no maxed star data found, fallback to empty (0 stars) data
                    print(f"No data available for '{item_name}' with {target_amber} amber and {target_cyan} cyan stars, falling back to empty sculpture")
                    empty_star_data = []
                    for entry in closed_data:
                        cyan_stars = entry.get('cyan_stars', 0)
                        amber_stars = entry.get('amber_stars', 0)
                        if cyan_stars == 0 and amber_stars == 0:
                            empty_star_data.append(entry)
                    
                    if empty_star_data:
                        closed_data = empty_star_data
                    else:
                        print(f"No empty sculpture data available for '{item_name}' either")
                        return None
            else:
                pass
                
        elif has_mod_rank_data and rank > 0:
            # For mods with specific rank requested
            rank_filtered_data = [
                entry for entry in closed_data 
                if 'mod_rank' in entry and entry['mod_rank'] == rank
            ]
            if not rank_filtered_data:
                print(f"No rank {rank} data available for mod '{item_name}'")
                return None
            closed_data = rank_filtered_data
        elif has_mod_rank_data and rank == 0:
            # For mods with rank 0 (unranked)
            unranked_data = [
                entry for entry in closed_data 
                if 'mod_rank' in entry and entry['mod_rank'] == 0
            ]
            if unranked_data:
                closed_data = unranked_data
            # If no rank 0 data, use all data (fallback)
        # For non-mod items (and non-ayatan sculptures), use all available data as-is

        if not closed_data:
            print(f"No suitable data found for '{item_name}' with rank {rank}")
            return None
        
        # Sort by datetime to get the most recent entry
        closed_data.sort(key=lambda x: x['datetime'], reverse=True)
        most_recent = closed_data[0]
        
        # Return the appropriate price based on method
        if price_method == 'minimum':
            return most_recent.get('min_price')
        else:  # median (default)
            return most_recent.get('median')
            
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            print(f"Item '{item_name}' not found in Warframe Market")
        else:
            print(f"HTTP Error {e.response.status_code} for '{item_name}': {e}")
        return None
    except Exception as e:
        print(f"Error fetching statistics for '{item_name}': {e}")
        return None


def get_item_price(item_name, price_method, rank):
    """
    Fetches the minimum sell price of the given item from the Warframe Market API.
    Supports rank filtering for mods.
    """
    url_name = normalize_item_name(item_name)
    api_url = f"https://api.warframe.market/v1/items/{url_name}/orders"

    headers = {
        'accept': 'application/json',
        'platform': 'pc',
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()
        data = response.json()

        orders = data['payload']['orders']
        
        # Apply rank filtering if needed
        def rank_filter(order):
            if rank == 0:
                # For rank 0, include unranked mods and non-mods
                return 'mod_rank' not in order or order['mod_rank'] == 0
            else:
                # For specific ranks, only include matching mods
                return 'mod_rank' in order and order['mod_rank'] == rank
        
        # For minimum prices
        if price_method == 'minimum':
            # Filter for online sellers matching rank
            online_orders = [
                order for order in orders 
                if order['order_type'] == 'sell' and 
                order['user']['status'] == 'ingame' and
                rank_filter(order)
            ]
            
            if online_orders:
                min_price = min(order['platinum'] for order in online_orders)
                return min_price
            else:
                # Fallback to all sell orders with rank filter
                all_sell_orders = [
                    order for order in orders 
                    if order['order_type'] == 'sell' and
                    rank_filter(order)
                ]
                if all_sell_orders:
                    min_price = min(order['platinum'] for order in all_sell_orders)
                    return min_price
                else:
                    return None
                    
        # For median prices
        else:
            # Collect all valid sell orders with quantity expansion
            all_prices = []
            for order in orders:
                if (
                    order['order_type'] == 'sell' and 
                    order['user']['status'] == 'ingame' and
                    rank_filter(order)):
                    
                    # Add price for each available item in the order
                    all_prices.extend([order['platinum']] * order['quantity'])
                
            if all_prices: #TODO add handling for filtered median to remove expensive outliers (useful for low traded mods)
                # Calculate true quantity-weighted median
                sorted_prices = sorted(all_prices)
                n = len(sorted_prices)
                
                if n % 2 == 1:
                    # Odd number of items
                    median_price = sorted_prices[n//2]
                else:
                    # Even number of items
                    median_price = (sorted_prices[n//2 - 1] + sorted_prices[n//2]) / 2
                
                return median_price
            else:
                # Fallback to minimum price among all sell orders
                sell_orders = [
                    order for order in orders 
                    if order['order_type'] == 'sell' and
                    rank_filter(order)
                ]
                if sell_orders:
                    min_price = min(order['platinum'] for order in sell_orders)
                    return min_price
                else:
                    return None
    except Exception as e:
        print(f"Error fetching price for '{item_name}': {e}")
        return None

def write_to_excel(items, output_file, price_method, api_method):
    """
    Writes item data to Excel with rank support
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warframe Market Prices"

    # Write headers
    headers = ['Quantity', 'Item', 'Rank', 'Item Value', 'Total Value']
    ws.append(headers)

    total_sum = 0

    for quantity, item_name, rank in items:
        if api_method == 'listings':
            price = get_item_price(item_name, price_method, rank)
        else:  # statistics (default)
            price = get_item_price_stat(item_name, price_method, rank)
        
        total_value = quantity * price if price is not None else None
        row = [quantity, item_name, rank, price, total_value]
        ws.append(row)

        # Accumulate the total if the value is available
        if total_value is not None:
            total_sum += total_value

    # Append the total row
    total_row = ['', 'Total', '', '', total_sum]
    ws.append(total_row)

    wb.save(output_file)
    print(f"Data written to {output_file}")

# Ayatan Sculpture maxed data: {normalized_name: (amber_stars, cyan_stars)}
ayatan_sculptures = {
    'ayatan_anasa_sculpture': (2, 2),
    'ayatan_hemakara_sculpture': (1, 2),
    'ayatan_kitha_sculpture': (1, 4),
    'ayatan_zambuka_sculpture': (1, 2),
    'ayatan_orta_sculpture': (1, 3),
    'ayatan_vaya_sculpture': (1, 2),
    'ayatan_piv_sculpture': (1, 2),
    'ayatan_valana_sculpture': (1, 2),
    'ayatan_sah_sculpture': (1, 2),
    'ayatan_ayr_sculpture': (0, 3),
}

# Weird URL special cases handler
special_cases = {
    "semi-shotgun cannonade": "shotgun_cannonade",
    "summoner's wrath": "summoner%E2%80%99s_wrath",
    "fear sense": "sense_danger",
    "negation armor": "negation_swarm",
    "teleport rush": "fatal_teleport",
    "ghoulsaw blade": "ghoulsaw_blade_blueprint",
    "ghoulsaw engine": "ghoulsaw_engine_blueprint",
    "ghoulsaw grip": "ghoulsaw_grip_blueprint",
    "mutalist alad v assassinate (key)": "mutalist_alad_v_assassinate_key",
    "mutalist alad v nav coordinate": "mutalist_nav_coordinates",
    "scan aquatic lifeforms": "scan_lifeforms",
    "orokin tower extraction scene": "orokin_tower_extraction_scene",
    "orokin derelict simulacrum": "orokin_derelict_simulacrum",
    "orokin derelict plaza scene": "orokin_derelict_plaza_scene",
    "central mall backroom scene": "central_mall_backroom",
    "höllvanian historic quarter in spring scene": "höllvanian_historic_quarter_in_spring",
    "höllvanian intersection in winter scene": "höllvanian_intersection_in_winter",
    "höllvanian old town in fall scene": "höllvanian_old_town_in_fall",
    "höllvanian tenements in summer scene": "höllvanian_tenements_in_summer",
    "höllvanian terrace in summer scene": "höllvanian_terrace_in_summer",
    "orbit arcade scene": "orbit_arcade",
    "tech titan electronics store scene": "tech_titan_electronics_store",
    "riot-848 stock blueprint": "riot_848_stock",
    "riot-848 barrel blueprint": "riot_848_barrel",
    "riot-848 receiver blueprint": "riot_848_receiver",
    }
exceptions = [
    'carrier_prime_systems',
    'dethcube_prime_systems',
    'helios_prime_systems',
    'nautilus_prime_systems',
    'nautilus_systems',
    'shade_prime_systems',
    'shedu_chassis',
    'spectra_vandal_chassis',
    'wyrm_prime_systems'
]

def main():
    parser = argparse.ArgumentParser(
        description='Warframe Market Price Calculator'
    )
    
    parser.add_argument('-i', '--input', 
                        default='input.txt',
                        help='Input file (TXT, CSV, or XLSX)')
    parser.add_argument('-o', '--output', 
                        default='output.xlsx',
                        help='Output Excel file')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('-l', '--listings', 
                        action='store_true',
                        help='Use listings (orders) API')
    group.add_argument('-s', '--statistics', 
                        action='store_true',
                        help='Use statistics API (default)')
    group2 = parser.add_mutually_exclusive_group()
    group2.add_argument('-m', '--minimum', 
                       action='store_true',
                       help='Use minimum prices')
    group2.add_argument('-M', '--median', 
                       action='store_true',
                       help='Use median prices (default)')
    
    args = parser.parse_args()
    
    price_method = 'minimum' if args.minimum else 'median'
    api_method = 'listings' if args.listings else 'statistics'
    
    try:
        print(f"Processing input file: {args.input}")
        items = parse_input(args.input)
        
        if api_method == 'listings':
            print(f"Found {len(items)} items, fetching {price_method} prices using orders API...")
        else:
            print(f"Found {len(items)} items, fetching {price_method} prices using statistics API...")
        
        write_to_excel(items, args.output, price_method, api_method)
        print("Operation completed successfully!")
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()